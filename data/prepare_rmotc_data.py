"""
Prepare RMOTC Teapot Dome (NPR-3) dataset.

Extracts rmotc.tar, indexes LAS logs, parses well headers / directional surveys,
copies seismic SEG-Y, writes metadata under data/rmotc/prepared/.
"""

from __future__ import annotations

import csv
import json
import re
import shutil
import tarfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np

from .prepare_volve_data import STANDARD_CURVES, annotate_geometry_quality, score_las_file

FT_TO_M = 0.3048


def normalize_api(val) -> str:
    """Normalize API to digit string without leading zeros (12–14 digit RMOTC IDs)."""
    digits = re.sub(r"\D", "", str(val))
    return digits.lstrip("0") or "0"


def api12(val) -> str:
    """RMOTC LAS filenames use 14-digit API = 12-digit API + '00' suffix."""
    d = normalize_api(val)
    if len(d) > 12:
        return d[:12]
    return d


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise ImportError("pip install openpyxl") from e


def _norm_col(name: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", str(name).strip().upper()).strip("_")


def _pick_column(cols: List[str], candidates: List[str]) -> Optional[str]:
    norm = {_norm_col(c): c for c in cols}
    for cand in candidates:
        key = _norm_col(cand)
        if key in norm:
            return norm[key]
    for cand in candidates:
        key = _norm_col(cand)
        for nk, orig in norm.items():
            if key in nk or nk in key:
                return orig
    return None


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def extract_rmotc_tar(tar_path: Path, dest_dir: Path, force: bool = False) -> Path:
    marker = dest_dir / "DataSets" / "Well Log"
    if marker.exists() and not force:
        print(f"RMOTC already extracted at {dest_dir}")
        return dest_dir
    dest_dir.mkdir(parents=True, exist_ok=True)
    print(f"Extracting {tar_path} -> {dest_dir} ...")
    with tarfile.open(tar_path, "r") as tf:
        members = [m for m in tf.getmembers() if "/._" not in m.name and not m.name.startswith("._")]
        tf.extractall(dest_dir, members=members)
    print("Extraction complete.")
    return dest_dir


def load_well_headers(xlsx_path: Path) -> Dict[str, Dict]:
    _ensure_openpyxl()
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    cols: List[str] = []
    rows_iter = ws.iter_rows(values_only=True)
    for row in rows_iter:
        if not row or row[0] is None:
            continue
        if str(row[0]).strip() == "API Number":
            cols = [str(c).strip() if c is not None else "" for c in row]
            break

    if not cols:
        wb.close()
        raise ValueError(f"No header row found in {xlsx_path}")

    api_col = _pick_column(cols, ["API_NUMBER", "API"])
    name_col = _pick_column(cols, ["WELL_NAME", "WELL"])
    north_col = _pick_column(cols, ["NORTHING", "NORTH"])
    east_col = _pick_column(cols, ["EASTING", "EAST"])
    kb_col = _pick_column(cols, ["DATUM_ELEVATION", "DATUM_ELEV", "KB", "ELEV"])
    idx = {c: i for i, c in enumerate(cols)}

    def g(row, col):
        if col is None:
            return None
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    out: Dict[str, Dict] = {}
    for row in rows_iter:
        if not row or row[0] is None:
            continue
        if str(row[0]).strip() == "API Number":
            continue
        api_raw = g(row, api_col)
        if api_raw is None:
            continue
        api = api12(api_raw)
        northing = _to_float(g(row, north_col))
        easting = _to_float(g(row, east_col))
        kb = _to_float(g(row, kb_col))
        if kb is not None and kb > 500:
            kb *= FT_TO_M
        name = g(row, name_col)
        wnum = g(row, _pick_column(cols, ["WELL_NUMBER", "WELL_NUM"]))
        well_label = str(name).strip() if name else api
        if wnum:
            well_label = f"{well_label}-{wnum}".strip("-")

        out[api] = {
            "api": api,
            "api_raw": str(api_raw),
            "well_name": well_label,
            "surface_y": northing,
            "surface_x": easting,
            "kb_elevation_m": kb,
            "coord_source": "rmotc_well_headers",
            "crs": "Wyoming East Central State Plane NAD27",
        }
    wb.close()
    return out


def _is_survey_header(row) -> bool:
    if not row or row[0] is None:
        return False
    c0 = re.sub(r"[^A-Z0-9]", "", str(row[0]).upper())
    c1 = str(row[1]).upper() if len(row) > 1 and row[1] is not None else ""
    return c0 in ("APINUMBER", "API", "API#") and "MD" in c1


def load_directional_surveys(xlsx_path: Path) -> Dict[str, List[Tuple]]:
    _ensure_openpyxl()
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    surveys: Dict[str, List] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        i = 0
        while i < len(rows):
            row = rows[i]
            if _is_survey_header(row):
                i += 1
                while i < len(rows):
                    dr = rows[i]
                    if not dr or all(v is None for v in dr):
                        i += 1
                        continue
                    label = str(dr[0]).strip()
                    if label.upper().replace(":", "").startswith("WELL") or _is_survey_header(dr):
                        break
                    api = api12(dr[0])
                    md = _to_float(dr[1])
                    inc = _to_float(dr[2])
                    if md is None or inc is None:
                        i += 1
                        continue
                    azi = _to_float(dr[3]) if len(dr) > 3 else 0.0
                    surveys.setdefault(api, []).append((
                        md * FT_TO_M, inc, azi or 0.0, None, None, None,
                    ))
                    i += 1
            else:
                i += 1

    wb.close()
    for api in surveys:
        surveys[api].sort(key=lambda p: p[0])
    return surveys


def discover_las_files(raw_root: Path) -> Dict[str, List[Path]]:
    las_root = raw_root / "DataSets" / "Well Log" / "CD Files" / "LAS_log_files"
    out: Dict[str, List[Path]] = {}
    for sub in ("Shallow_LAS_files", "Deeper_LAS_files"):
        base = las_root / sub
        if not base.exists():
            continue
        for p in list(base.rglob("*.las")) + list(base.rglob("*.LAS")):
            if p.name.startswith("._"):
                continue
            m = re.match(r"(\d{10,14})", p.stem)
            if m:
                out.setdefault(api12(m.group(1)), []).append(p)
    return out


def pick_best_las(paths: List[Path]) -> Optional[Path]:
    best, best_score = None, -1
    for p in paths:
        try:
            score, _, _ = score_las_file(p)
            if score > best_score:
                best_score, best = score, p
        except Exception:
            pass
    return best


def write_vertical_deviation(out_path: Path, td_m: float = 3000.0, step: float = 10.0):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["MD", "INCL", "AZI", "TVD", "NS", "EW", "SOURCE"])
        md = 0.0
        while md <= td_m:
            w.writerow([f"{md:.2f}", "0.00", "0.00", f"{md:.2f}", "0.00", "0.00", "vertical_assumed"])
            md += step


def write_survey_csv(survey: List[Tuple], out_path: Path, source: str):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["MD", "INCL", "AZI", "TVD", "NS", "EW", "SOURCE"])
        for md, inc, azi, tvd, ns, ew in survey:
            if tvd is None:
                tvd = md * np.cos(np.radians(inc))
            if ns is None:
                ns = md * np.sin(np.radians(inc)) * np.cos(np.radians(azi))
            if ew is None:
                ew = md * np.sin(np.radians(inc)) * np.sin(np.radians(azi))
            w.writerow([f"{md:.2f}", f"{inc:.2f}", f"{azi:.2f}", f"{tvd:.2f}", f"{ns:.2f}", f"{ew:.2f}", source])


def validate_segy(segy_path: Path) -> Dict:
    import segyio
    from segyio import BinField, TraceField

    with segyio.open(str(segy_path), "r", ignore_geometry=True) as f:
        tr0 = f.trace[0].astype(np.float64)
        return {
            "ok": True,
            "tracecount": f.tracecount,
            "nsamples": len(f.samples),
            "sample_interval_us": int(f.bin[BinField.Interval]),
            "trace0_std": float(np.std(tr0)),
        }


def setup_seismic(raw_root: Path, seismic_dir: Path) -> Dict:
    seismic_dir.mkdir(parents=True, exist_ok=True)
    out: Dict = {}
    src_3d = raw_root / "DataSets/Seismic/CD files/3D_Seismic/filt_mig.sgy"
    if src_3d.exists():
        dst = seismic_dir / "filt_mig.segy"
        if not dst.exists() or dst.stat().st_size != src_3d.stat().st_size:
            shutil.copy2(src_3d, dst)
        out["segy_3d"] = str(dst)
    src_2d = raw_root / "DataSets/Seismic/CD files/2D_Seismic/NormalizedMigrated_segy"
    if src_2d.exists():
        out["segy_2d"] = {}
        for src in sorted(src_2d.glob("line*.sgy")):
            if src.name.startswith("._"):
                continue
            dst = seismic_dir / "2d" / src.name.replace(".sgy", ".segy")
            dst.parent.mkdir(parents=True, exist_ok=True)
            if not dst.exists() or dst.stat().st_size != src.stat().st_size:
                shutil.copy2(src, dst)
            out["segy_2d"][src.stem] = str(dst)
    return out


def prepare_rmotc_data(project_root: Optional[Path] = None, tar_path: Optional[Path] = None, force_extract: bool = False) -> Dict:
    root = Path(project_root or Path(__file__).resolve().parents[1])
    tar_path = Path(tar_path or root / "data" / "rmotc.tar")
    raw_dir = root / "data" / "rmotc" / "raw"
    prepared = root / "data" / "rmotc" / "prepared"
    dev_dir = prepared / "deviations"
    seismic_dir = root / "data" / "rmotc" / "seismic"
    prepared.mkdir(parents=True, exist_ok=True)

    extract_rmotc_tar(tar_path, raw_dir, force=force_extract)
    cd = raw_dir / "DataSets/Well Log/CD Files"
    headers = load_well_headers(cd / "TeapotDomeWellHeaders02-09-10.xlsx")
    surveys = load_directional_surveys(cd / "DirectionalSurveys_020910.xlsx") if (cd / "DirectionalSurveys_020910.xlsx").exists() else {}
    las_map = discover_las_files(raw_dir)
    seismic_paths = setup_seismic(raw_dir, seismic_dir)

    well_metadata, curve_inventory, deviation_info = {}, {}, {}
    for api in sorted(las_map):
        best = pick_best_las(las_map[api])
        if best is None:
            continue
        score, std_curves, las_hdr = score_las_file(best)
        hdr = headers.get(api, {})
        wname = hdr.get("well_name") or api
        well_metadata[wname] = {
            "well_name": wname, "api": api, "api_raw": hdr.get("api_raw"), "las_path": str(best),
            "standard_curves": std_curves, "n_standard_curves": len(std_curves),
            "las_score": score,
            "surface_x": hdr.get("surface_x"),
            "surface_y": hdr.get("surface_y"),
            "kb_elevation_m": hdr.get("kb_elevation_m") or las_hdr.get("kb_elevation_m"),
            "coord_source": hdr.get("coord_source", "las_header"),
            "crs": hdr.get("crs"),
            "dataset": "rmotc",
        }
        curve_inventory[wname] = {"curves": std_curves, "missing": [c for c in STANDARD_CURVES if c not in std_curves], "las_file": best.name, "api": api}
        csv_path = dev_dir / f"{api}.csv"
        if api in surveys:
            write_survey_csv(surveys[api], csv_path, "rmotc_directional")
            deviation_info[wname] = {"source": "rmotc_directional", "points": len(surveys[api]), "api": api}
        else:
            write_vertical_deviation(csv_path)
            deviation_info[wname] = {"source": "vertical_assumed", "points": 301, "api": api}

    annotate_geometry_quality(well_metadata, deviation_info)
    segy_val = validate_segy(Path(seismic_paths["segy_3d"])) if seismic_paths.get("segy_3d") else {}
    layout = {"dataset": "rmotc", "raw_dir": str(raw_dir), "prepared_dir": str(prepared),
              "seismic_dir": str(seismic_dir), **seismic_paths}

    for p, obj in [(prepared / "well_metadata.json", well_metadata), (prepared / "curve_inventory.json", curve_inventory),
                   (prepared / "deviation_inventory.json", deviation_info), (prepared / "data_layout.json", layout)]:
        with open(p, "w", encoding="utf-8") as f:
            json.dump(obj, f, indent=2, ensure_ascii=False)

    manifest = {
        "prepared_at": datetime.now(timezone.utc).isoformat(),
        "dataset": "rmotc",
        "n_wells_with_las": len(well_metadata),
        "n_with_directional": sum(1 for d in deviation_info.values() if d.get("source") == "rmotc_directional"),
        "n_geometry_verified": sum(1 for m in well_metadata.values() if m.get("geometry_verified")),
        "segy_validation": segy_val,
        "paths": layout,
    }
    with open(prepared / "manifest.json", "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    print(f"RMOTC: {len(well_metadata)} wells, verified={manifest['n_geometry_verified']}, segy traces={segy_val.get('tracecount')}")
    return manifest


if __name__ == "__main__":
    prepare_rmotc_data()
